"""Tests for source_data_cross_sheet tool."""

from __future__ import annotations

import json
from pathlib import Path

from engine.static_audit.tools.source_data_cross_sheet import (
    CrossSheetFinding,
    SheetColumn,
    extract_numeric_columns,
    find_cross_sheet_duplicates,
    run_cross_sheet_detection,
)


def test_cross_sheet_finding_to_dict() -> None:
    """Test CrossSheetFinding serialization."""
    finding = CrossSheetFinding(
        finding_id="CSD-0001",
        workbook_1="wb1.xlsx",
        sheet_1="Sheet1",
        column_1="A",
        column_1_label="Values",
        workbook_2="wb2.xlsx",
        sheet_2="Sheet2",
        column_2="B",
        column_2_label="Numbers",
        overlap_rows=20,
        equal_rows=18,
        support_rate=0.9,
    )

    data = finding.to_dict()

    assert data["finding_id"] == "CSD-0001"
    assert data["category"] == "cross_sheet_duplicate_columns"
    assert data["issue_category"] == "consistency"
    assert data["overlap_rows"] == 20
    assert data["equal_rows"] == 18
    assert data["support_rate"] == 0.9


def test_extract_numeric_columns_with_openpyxl(tmp_path: Path) -> None:
    """Test numeric column extraction from XLSX file."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Add header
        ws.append(["ID", "Value1", "Value2", "Text"])

        # Add data rows
        for i in range(1, 11):
            ws.append([i, float(i) * 1.5, float(i) * 2.0, f"row_{i}"])

        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        columns = extract_numeric_columns(xlsx_path, "Sheet1")

        # Should extract ID, Value1, Value2 (3 numeric columns)
        assert len(columns) == 3

        # Check that Value1 column has correct values
        value1_col = [c for c in columns if c.column_label == "Value1"][0]
        assert len(value1_col.values) == 10
        assert value1_col.values[0] == 1.5
        assert value1_col.values[9] == 15.0

    except ImportError:
        # Skip test if openpyxl not available
        pass


def test_find_cross_sheet_duplicates_identifies_duplicates(tmp_path: Path) -> None:
    """Test that cross-sheet duplicate detection finds duplicates."""
    try:
        from openpyxl import Workbook

        # Create two XLSX files with overlapping numeric columns
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sheet1"
        ws1.append(["ID", "Value"])
        for i in range(1, 21):
            ws1.append([i, float(i) * 2.0])
        wb1.save(tmp_path / "file1.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet2"
        ws2.append(["ID", "Value"])
        # Same values as file1 for first 15 rows
        for i in range(1, 16):
            ws2.append([i, float(i) * 2.0])
        # Different values for remaining rows
        for i in range(16, 21):
            ws2.append([i, float(i) * 3.0])
        wb2.save(tmp_path / "file2.xlsx")

        findings = find_cross_sheet_duplicates(
            tmp_path,
            min_overlap=10,
            min_support_rate=0.7,
            max_findings=10,
        )

        # Should find at least one duplicate (ID columns match perfectly)
        assert len(findings) >= 1

        # Check that Value columns are detected as duplicates (15/20 = 0.75 support)
        value_findings = [f for f in findings if "Value" in f.column_1_label or "Value" in f.column_2_label]
        if value_findings:
            # Support rate should be around 0.75 (15 matching out of 20 overlapping)
            assert 0.7 <= value_findings[0].support_rate <= 0.8

    except ImportError:
        # Skip test if openpyxl not available
        pass


def test_find_cross_sheet_duplicates_respects_min_overlap(tmp_path: Path) -> None:
    """Test that min_overlap parameter is respected."""
    try:
        from openpyxl import Workbook

        # Create files with only 5 overlapping rows
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.append(["Value"])
        for i in range(1, 6):
            ws1.append([float(i)])
        wb1.save(tmp_path / "file1.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Value"])
        for i in range(1, 6):
            ws2.append([float(i)])
        wb2.save(tmp_path / "file2.xlsx")

        # With min_overlap=10, should find no duplicates
        findings = find_cross_sheet_duplicates(
            tmp_path,
            min_overlap=10,
            min_support_rate=0.8,
            max_findings=10,
        )
        assert len(findings) == 0

        # With min_overlap=5, should find duplicates
        findings = find_cross_sheet_duplicates(
            tmp_path,
            min_overlap=5,
            min_support_rate=0.8,
            max_findings=10,
        )
        assert len(findings) >= 1

    except ImportError:
        pass


def test_find_cross_sheet_duplicates_respects_min_support_rate(tmp_path: Path) -> None:
    """Test that min_support_rate parameter is respected."""
    try:
        from openpyxl import Workbook

        # Create files with 50% matching values
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.append(["Value"])
        for i in range(1, 21):
            ws1.append([float(i)])
        wb1.save(tmp_path / "file1.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["Value"])
        # First 10 rows match, next 10 don't
        for i in range(1, 11):
            ws2.append([float(i)])
        for i in range(11, 21):
            ws2.append([float(i) * 10.0])  # Different values
        wb2.save(tmp_path / "file2.xlsx")

        # With min_support_rate=0.8, should find no duplicates (only 50% match)
        findings = find_cross_sheet_duplicates(
            tmp_path,
            min_overlap=10,
            min_support_rate=0.8,
            max_findings=10,
        )
        assert len(findings) == 0

        # With min_support_rate=0.4, should find duplicates
        findings = find_cross_sheet_duplicates(
            tmp_path,
            min_overlap=10,
            min_support_rate=0.4,
            max_findings=10,
        )
        assert len(findings) >= 1

    except ImportError:
        pass


def test_run_cross_sheet_detection_output_format(tmp_path: Path) -> None:
    """Test that run_cross_sheet_detection returns correct format."""
    try:
        from openpyxl import Workbook

        # Create a simple XLSX file
        wb = Workbook()
        ws = wb.active
        ws.append(["Value"])
        for i in range(1, 11):
            ws.append([float(i)])
        wb.save(tmp_path / "test.xlsx")

        result = run_cross_sheet_detection(
            tmp_path,
            min_overlap=5,
            min_support_rate=0.8,
            max_findings=10,
        )

        # Check output structure
        assert "findings" in result
        assert "parameters" in result
        assert result["parameters"]["min_overlap"] == 5
        assert result["parameters"]["min_support_rate"] == 0.8
        assert result["parameters"]["max_findings"] == 10

        # findings should be a list
        assert isinstance(result["findings"], list)

    except ImportError:
        pass


def test_run_cross_sheet_detection_with_no_xlsx_files(tmp_path: Path) -> None:
    """Test detection with directory containing no XLSX files."""
    result = run_cross_sheet_detection(
        tmp_path,
        min_overlap=10,
        min_support_rate=0.8,
        max_findings=10,
    )

    assert result["findings"] == []
    assert "parameters" in result


def test_sheet_column_dataclass() -> None:
    """Test SheetColumn dataclass."""
    col = SheetColumn(
        workbook="test.xlsx",
        sheet="Sheet1",
        column="A",
        column_label="Value",
        values=[10.0, 20.0, 30.0],
        row_indices=[1, 2, 3],
    )

    assert col.workbook == "test.xlsx"
    assert col.sheet == "Sheet1"
    assert col.column == "A"
    assert col.column_label == "Value"
    assert len(col.values) == 3
    assert len(col.row_indices) == 3
