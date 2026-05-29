#!/usr/bin/env python3
"""Render annotated PNG evidence cards from source-data audit findings.

The script is deterministic: it reads the original XLSX cells and an audit JSON,
then redraws the relevant table region with highlighted anomalous ranges.
"""

import argparse
import json
import math
import re
from pathlib import Path
from textwrap import wrap

import openpyxl
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from PIL import Image, ImageDraw, ImageFont


PALETTE = {
    "bg": "#f7f8fb",
    "paper": "#ffffff",
    "grid": "#d7dce5",
    "header": "#eef2f7",
    "text": "#18202f",
    "muted": "#5f6b7a",
    "danger": "#d92d20",
    "danger_fill": "#fff1f0",
    "warn": "#f79009",
    "note": "#1d4ed8",
}


def load_font(size, bold=False):
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size, index=1 if bold else 0)
        except Exception:
            continue
    return ImageFont.load_default()


FONT = {
    "title": load_font(26, True),
    "subtitle": load_font(17),
    "body": load_font(15),
    "small": load_font(13),
    "mono": load_font(14),
}


def parse_a1_range(text):
    m = re.search(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", text)
    if not m:
        raise ValueError(f"Cannot parse A1 range from: {text}")
    c1, r1, c2, r2 = m.groups()
    return int(r1), column_index_from_string(c1), int(r2), column_index_from_string(c2)


def all_ranges(text):
    ranges = []
    for m in re.finditer(r"([A-Z]+\d+:[A-Z]+\d+)", text):
        ranges.append(parse_a1_range(m.group(1)))
    if not ranges:
        ranges.append(parse_a1_range(text))
    return ranges


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if math.isfinite(v):
            return f"{v:.10g}"
        return str(v)
    return str(v)


def terminal_digit(v):
    s = fmt(v).strip()
    if not s:
        return None
    if "e" in s.lower():
        try:
            s = f"{float(s):.12f}".rstrip("0").rstrip(".")
        except Exception:
            pass
    for ch in reversed(s):
        if ch.isdigit():
            return ch
    return None


def text_size(draw, text, font):
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0], box[3] - box[1]


def draw_wrapped(draw, text, xy, width, font, fill, line_gap=5):
    x, y = xy
    # Wrap by rough character count; works for mixed CN/EN without measuring each word.
    chars = max(10, int(width / 8))
    lines = []
    for para in str(text).splitlines() or [""]:
        lines.extend(wrap(para, chars) or [""])
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        _, h = text_size(draw, line or "Ag", font)
        y += h + line_gap
    return y


def crop_window(finding, max_rows):
    ranges = all_ranges(finding["range"])
    r1 = min(r[0] for r in ranges)
    c1 = min(r[1] for r in ranges)
    r2 = max(r[2] for r in ranges)
    c2 = max(r[3] for r in ranges)
    r1 = max(1, r1 - 3)
    c1 = max(1, c1 - 1)
    r2 = r2 + 1
    c2 = c2 + 1
    if r2 - r1 + 1 > max_rows:
        r2 = r1 + max_rows - 1
    return r1, c1, r2, c2


def highlighted_cells(finding, window):
    r1, c1, r2, c2 = window
    cells = set()
    for a, b, c, d in all_ranges(finding["range"]):
        for r in range(max(r1, a), min(r2, c) + 1):
            for col in range(max(c1, b), min(c2, d) + 1):
                cells.add((r, col))
    return cells


def evidence_note(finding):
    if "constant_difference_col1_minus_col2" in finding:
        return (
            f"Fixed difference: n={finding.get('n')}, "
            f"col1-col2={finding.get('constant_difference_col1_minus_col2')}"
        )
    if "top_digit" in finding:
        return (
            f"Terminal digit concentration: n={finding.get('n')}, "
            f"digit {finding.get('top_digit')}={finding.get('top_count')} "
            f"({finding.get('top_ratio')})"
        )
    return finding.get("severity", "High-risk finding")


def render_finding(finding, xlsx_root, output_dir, index, max_rows=42):
    workbook = xlsx_root / finding["file"]
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    ws = wb[finding["sheet"]]

    window = crop_window(finding, max_rows=max_rows)
    r1, c1, r2, c2 = window
    highlights = highlighted_cells(finding, window)

    rows = list(range(r1, r2 + 1))
    cols = list(range(c1, c2 + 1))
    cell_w = 98
    cell_h = 34
    row_head_w = 58
    col_head_h = 32
    left = 44
    top = 130
    table_w = row_head_w + len(cols) * cell_w
    table_h = col_head_h + len(rows) * cell_h
    note_w = 360
    margin = 36
    img_w = left + table_w + note_w + margin * 2
    img_h = max(top + table_h + 50, 720)

    img = Image.new("RGB", (img_w, img_h), PALETTE["bg"])
    draw = ImageDraw.Draw(img)

    title = f"High-risk Evidence #{index}: {finding['file']} / {finding['sheet']}"
    draw.text((left, 32), title, font=FONT["title"], fill=PALETTE["text"])
    subtitle = f"Range: {finding['range']}"
    draw.text((left, 70), subtitle, font=FONT["subtitle"], fill=PALETTE["muted"])

    # Table background.
    draw.rounded_rectangle(
        (left, top, left + table_w, top + table_h),
        radius=8,
        fill=PALETTE["paper"],
        outline=PALETTE["grid"],
        width=1,
    )

    # Column headers.
    draw.rectangle((left, top, left + table_w, top + col_head_h), fill=PALETTE["header"])
    draw.rectangle((left, top, left + row_head_w, top + table_h), fill=PALETTE["header"])
    for j, col in enumerate(cols):
        x = left + row_head_w + j * cell_w
        draw.rectangle((x, top, x + cell_w, top + col_head_h), outline=PALETTE["grid"])
        label = get_column_letter(col)
        tw, th = text_size(draw, label, FONT["small"])
        draw.text((x + (cell_w - tw) / 2, top + 8), label, font=FONT["small"], fill=PALETTE["muted"])

    # Rows and cells.
    for i, row in enumerate(rows):
        y = top + col_head_h + i * cell_h
        draw.rectangle((left, y, left + row_head_w, y + cell_h), outline=PALETTE["grid"], fill=PALETTE["header"])
        label = str(row)
        tw, th = text_size(draw, label, FONT["small"])
        draw.text((left + row_head_w - tw - 10, y + 9), label, font=FONT["small"], fill=PALETTE["muted"])
        for j, col in enumerate(cols):
            x = left + row_head_w + j * cell_w
            fill = PALETTE["danger_fill"] if (row, col) in highlights else PALETTE["paper"]
            draw.rectangle((x, y, x + cell_w, y + cell_h), fill=fill, outline=PALETTE["grid"])
            value = fmt(ws.cell(row, col).value)
            if len(value) > 12:
                value = value[:11] + "…"
            color = PALETTE["danger"] if (row, col) in highlights else PALETTE["text"]
            draw.text((x + 7, y + 9), value, font=FONT["mono"], fill=color)

    # Strong outline around highlighted ranges.
    for a, b, c, d in all_ranges(finding["range"]):
        aa, bb, cc, dd = max(r1, a), max(c1, b), min(r2, c), min(c2, d)
        if aa > cc or bb > dd:
            continue
        x0 = left + row_head_w + (bb - c1) * cell_w
        y0 = top + col_head_h + (aa - r1) * cell_h
        x1 = left + row_head_w + (dd - c1 + 1) * cell_w
        y1 = top + col_head_h + (cc - r1 + 1) * cell_h
        draw.rectangle((x0, y0, x1, y1), outline=PALETTE["danger"], width=4)

    # Annotation panel.
    note_x = left + table_w + 30
    note_y = top
    draw.rounded_rectangle((note_x, note_y, note_x + note_w, note_y + 360), radius=8, fill=PALETTE["paper"], outline=PALETTE["grid"])
    draw.text((note_x + 20, note_y + 20), "Risk annotation", font=FONT["subtitle"], fill=PALETTE["danger"])
    y = note_y + 58
    y = draw_wrapped(draw, evidence_note(finding), (note_x + 20, y), note_w - 40, FONT["body"], PALETTE["text"])
    y += 12
    header = finding.get("header") or finding.get("header_col1") or ""
    if finding.get("header_col2"):
        header = f"{header} / {finding.get('header_col2')}"
    y = draw_wrapped(draw, f"Header: {header or 'not identified'}", (note_x + 20, y), note_w - 40, FONT["small"], PALETTE["muted"])
    y += 12
    y = draw_wrapped(draw, "Review note: This image is generated from the original XLSX cells. It is an evidence locator, not a standalone fraud verdict.", (note_x + 20, y), note_w - 40, FONT["small"], PALETTE["muted"])

    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", f"{index:02d}_{finding['file']}_{finding['sheet']}_{finding['range']}")
    path = output_dir / f"{safe[:180]}.png"
    img.save(path)
    return path


def main():
    parser = argparse.ArgumentParser(description="Render annotated source-data evidence PNGs.")
    parser.add_argument("--audit-json", required=True, help="Audit JSON containing terminal_findings/fixed_difference_findings.")
    parser.add_argument("--xlsx-root", required=True, help="Folder containing the source-data XLSX files.")
    parser.add_argument("--output", required=True, help="Output folder for PNG images.")
    parser.add_argument("--top-fixed", type=int, default=3)
    parser.add_argument("--top-terminal", type=int, default=8)
    parser.add_argument("--max-rows", type=int, default=42)
    args = parser.parse_args()

    audit = json.loads(Path(args.audit_json).read_text(encoding="utf-8"))
    xlsx_root = Path(args.xlsx_root)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    findings = []
    findings.extend(audit.get("fixed_difference_findings", [])[: args.top_fixed])
    findings.extend(audit.get("terminal_findings", [])[: args.top_terminal])

    manifest = []
    for i, finding in enumerate(findings, 1):
        path = render_finding(finding, xlsx_root, output_dir, i, max_rows=args.max_rows)
        manifest.append({
            "index": i,
            "image": str(path),
            "file": finding.get("file"),
            "sheet": finding.get("sheet"),
            "range": finding.get("range"),
            "note": evidence_note(finding),
        })

    manifest_path = output_dir / "evidence_images_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"count": len(manifest), "manifest": str(manifest_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
