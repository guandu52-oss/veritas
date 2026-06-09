"""Cross-sheet duplicate column detection for Source Data.

Detects when the same numeric column appears in multiple sheets or workbooks,
which may indicate data fabrication (e.g., same control group appearing in
multiple experiments, or same data packaged as different experiments).
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass, field
from itertools import combinations
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]


@dataclass
class SheetColumn:
    workbook: str
    sheet: str
    column: str
    column_label: str
    values: list[float | int]
    row_indices: list[int]


@dataclass
class CrossSheetFinding:
    finding_id: str
    workbook_1: str
    sheet_1: str
    column_1: str
    column_1_label: str
    workbook_2: str
    sheet_2: str
    column_2: str
    column_2_label: str
    overlap_rows: int
    equal_rows: int
    support_rate: float
    benign_explanations: list[str] = field(default_factory=list)
    pressure_test_result: str = "needs_review"
    manual_review_note: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "finding_id": self.finding_id,
            "category": "cross_sheet_duplicate_columns",
            "issue_category": "consistency",
            "workbook_1": self.workbook_1,
            "sheet_1": self.sheet_1,
            "column_1": self.column_1,
            "column_1_label": self.column_1_label,
            "workbook_2": self.workbook_2,
            "sheet_2": self.sheet_2,
            "column_2": self.column_2,
            "column_2_label": self.column_2_label,
            "overlap_rows": self.overlap_rows,
            "equal_rows": self.equal_rows,
            "support_rate": self.support_rate,
            "benign_explanations": self.benign_explanations,
            "pressure_test_result": self.pressure_test_result,
            "manual_review_note": self.manual_review_note,
        }


def extract_numeric_columns(
    workbook_path: Path,
    sheet_name: str,
    min_numeric_ratio: float = 0.5,
) -> list[SheetColumn]:
    """Extract numeric columns from a sheet.

    Returns list of SheetColumn where at least min_numeric_ratio of non-empty
    cells contain numeric values.
    """
    if load_workbook is None:
        return []

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    columns: list[SheetColumn] = []
    for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
        # Skip header row
        data = list(col[1:]) if len(col) > 1 else []
        if not data:
            continue

        # Count numeric values
        numeric_values = []
        row_indices = []
        for row_idx, value in enumerate(data, start=2):
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric_values.append(value)
                row_indices.append(row_idx)

        # Check if column is mostly numeric
        non_empty = [v for v in data if v is not None and v != ""]
        if len(non_empty) == 0:
            continue

        numeric_ratio = len(numeric_values) / len(non_empty)
        if numeric_ratio < min_numeric_ratio:
            continue

        # Get column label from header
        col_letter = ws.cell(row=1, column=col_idx).value or f"Col_{col_idx}"

        columns.append(
            SheetColumn(
                workbook=workbook_path.name,
                sheet=sheet_name,
                column=col_letter,
                column_label=str(col_letter),
                values=numeric_values,
                row_indices=row_indices,
            )
        )

    wb.close()
    return columns


def find_cross_sheet_duplicates(
    source_data_dir: Path,
    min_overlap: int = 10,
    min_support_rate: float = 0.8,
    max_findings: int = 50,
) -> list[CrossSheetFinding]:
    """Find duplicate columns across different sheets/workbooks.

    Args:
        source_data_dir: Directory containing XLSX files
        min_overlap: Minimum number of overlapping rows to consider
        min_support_rate: Minimum ratio of equal values (equal_rows / overlap_rows)
        max_findings: Maximum number of findings to return

    Returns:
        List of CrossSheetFinding sorted by support_rate (descending)
    """
    if load_workbook is None:
        return []

    # Collect all numeric columns from all sheets in all workbooks
    all_columns: list[SheetColumn] = []

    for xlsx_file in sorted(source_data_dir.glob("*.xlsx")):
        if xlsx_file.name.startswith("~$"):
            continue

        try:
            wb = load_workbook(xlsx_file, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            for sheet_name in sheet_names:
                columns = extract_numeric_columns(xlsx_file, sheet_name)
                all_columns.extend(columns)
        except Exception:
            continue

    # Compare every pair of columns from different sheets
    findings: list[CrossSheetFinding] = []
    finding_counter = 0

    for col1, col2 in combinations(all_columns, 2):
        # Skip if same sheet
        if col1.workbook == col2.workbook and col1.sheet == col2.sheet:
            continue

        # Find overlapping rows
        rows1 = set(col1.row_indices)
        rows2 = set(col2.row_indices)
        overlap_rows = rows1 & rows2

        if len(overlap_rows) < min_overlap:
            continue

        # Compare values at overlapping rows
        values1 = dict(zip(col1.row_indices, col1.values))
        values2 = dict(zip(col2.row_indices, col2.values))

        equal_count = 0
        for row in overlap_rows:
            if abs(values1[row] - values2[row]) < 1e-9:
                equal_count += 1

        support_rate = equal_count / len(overlap_rows)
        if support_rate < min_support_rate:
            continue

        finding_counter += 1
        finding = CrossSheetFinding(
            finding_id=f"CSD-{finding_counter:04d}",
            workbook_1=col1.workbook,
            sheet_1=col1.sheet,
            column_1=col1.column,
            column_1_label=col1.column_label,
            workbook_2=col2.workbook,
            sheet_2=col2.sheet,
            column_2=col2.column,
            column_2_label=col2.column_label,
            overlap_rows=len(overlap_rows),
            equal_rows=equal_count,
            support_rate=support_rate,
            benign_explanations=[
                "可能是同一个对照组在多个实验中使用",
                "可能是数据整理时的复制粘贴错误",
                "需要人工确认两个 sheet 是否代表独立实验",
            ],
            manual_review_note=f"检查 {col1.workbook}/{col1.sheet}/{col1.column} 和 {col2.workbook}/{col2.sheet}/{col2.column} 是否代表不同的实验条件",
        )
        findings.append(finding)

        if len(findings) >= max_findings:
            break

    # Sort by support_rate descending
    findings.sort(key=lambda f: f.support_rate, reverse=True)
    return findings


def run_cross_sheet_detection(
    source_data_dir: Path,
    min_overlap: int = 10,
    min_support_rate: float = 0.8,
    max_findings: int = 50,
) -> dict[str, Any]:
    """Run cross-sheet duplicate detection and return results.

    Returns:
        Dictionary with findings and metadata
    """
    findings = find_cross_sheet_duplicates(
        source_data_dir,
        min_overlap=min_overlap,
        min_support_rate=min_support_rate,
        max_findings=max_findings,
    )

    return {
        "schema_version": "1.0",
        "tool_id": "source_data.cross_sheet",
        "source_data_dir": str(source_data_dir),
        "parameters": {
            "min_overlap": min_overlap,
            "min_support_rate": min_support_rate,
            "max_findings": max_findings,
        },
        "finding_count": len(findings),
        "findings": [f.to_dict() for f in findings],
        "limitations": [
            "仅检测数值列的跨 sheet 重复",
            "不检测行级别的跨 sheet 重复",
            "不支持非 XLSX 格式",
        ],
    }


def main() -> int:
    """CLI entry point for cross-sheet detection."""
    if len(sys.argv) < 2:
        print("Usage: python source_data_cross_sheet.py <source_data_dir> [output.json]", file=sys.stderr)
        return 1

    source_data_dir = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if not source_data_dir.exists():
        print(f"Error: {source_data_dir} does not exist", file=sys.stderr)
        return 1

    results = run_cross_sheet_detection(source_data_dir)

    if output_path:
        output_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Results written to {output_path}", file=sys.stderr)
    else:
        print(json.dumps(results, indent=2, ensure_ascii=False))

    return 0


if __name__ == "__main__":
    sys.exit(main())
